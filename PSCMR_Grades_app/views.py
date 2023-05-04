from math import ceil
from django.shortcuts import render,redirect
import openpyxl
import mimetypes
from django.http import HttpResponse
import os


def grade(a):
    if (a=='AB' or a<40):
        return "F"
    elif a>=40 and a<50:
        return "D"
    elif a>=50 and a<60:
        return "C"
    elif a>=60 and a<70:
        return "B"
    elif a>=70 and a<80:
        return "A"
    elif a>=80 and a<90:
        return "S"
    elif a>=90 and a<=100:
        return "O"

def gradeR20(a):
    if (a=='AB' or a<40):
        return "F"
    elif a>=40 and a<50:
        return "E"
    elif a>=50 and a<60:
        return "D"
    elif a>=60 and a<70:
        return "C"
    elif a>=70 and a<80:
        return "B"
    elif a>=80 and a<90:
        return "A"
    elif a>=90 and a<=100:
        return "A+"

def mainPO(n):
    if n>=80:
        return 3
    elif n>=70 and n<80:
        return 2
    elif n>=60 and n<70:
        return 1
    else:
        return 0

def gradepoint(a):
    if a=='AB': return 'AB'
    elif a=='O': return 10
    elif a=='S': return 9
    elif a=='A': return 8
    elif a=='B': return 7
    elif a=='C': return 6
    elif a=='D': return 5
    elif a=='F': return 0

def gradepointR20(a):
    if a=='AB': return 'AB'
    elif a=='A+': return 10
    elif a=='A': return 9
    elif a=='B': return 8
    elif a=='C': return 7
    elif a=='D': return 6
    elif a=='E': return 5
    elif a=='F': return 0

def none_or_not(x):
    if x==None or x=='None' or x=='': return False
    else: return True

# Create your views here.
def index(request):
    if "GET" == request.method:
        return render(request, 'index.html', {})
    else:
        Name=request.POST['FName']
        Branch=request.POST['branch']
        Year=request.POST['year']
        Semester=request.POST['semester']
        Regulation=request.POST['regulation']
        Academic=request.POST['academic']
        Course=request.POST['Course']
        Coursecode=request.POST['Coursecode']
        details = list()
        details = [Name,Branch,Year,Semester,Regulation,Academic,Course,Coursecode]

        depart = {'CSE':'Computer Science and Engineering','CSE-AI&DS':'Computer Science and Engineering','CSE-AI&ML':'Computer Science and Engineering','CSE-AI':'Computer Science and Engineering','CSE-DS':'Computer Science and Engineering','CSE-IoT&CS(BCT)':'Computer Science and Engineering','CSE-IoT':'Computer Science and Engineering','ECE':'Electronics and Communication Engineering','EEE':'Electrical and Electronics Engineering'}

        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet1 = wb["Marks"]
        # worksheet2 = wb["Mid Attainment"]
        # worksheet3 = wb["SEM_END_Evaluation"]
        worksheet4 = wb["Course_End_Survey(Indirect)"]
        worksheet5 = wb["PO ATTAINMENT"]
        # worksheet6 = wb["CO-PO_MAPPING"]

        if Regulation == "R19":


            #print(worksheet)
            Main_excel_sheet = {}
            excel_data = list()
            excel_data1 = list()
            excel_data2 = list()
            excel_data3 = list()
            excel_data4 = list()
            excel_data5 = list()
            excel_data6 = list()
            count,i = 0,0
            nums = ['0','1','2','3','4','5','6','7','8','9']
            # iterating over the rows and
            # getting value from each cell in row
            for row in worksheet1.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

            for row in excel_data:
                excel_data1.append(row[0:8])
                excel_data2.append(row[0:2])
                excel_data2[i].extend(row[8:16])
                excel_data3.append(row[0:2])
                excel_data3[i].extend(row[-1])
                i+=1
            
            list01 = ['Total Assignment - 1',"CO1 (Grade)","CO2 (Grade)","CO3 (Grade)",'Total Assignment - 2',"CO4 (Grade)","CO5 (Grade)","CO6 (Grade)"]
            excel_data1[0].extend(list01)
        

            Total_Students = len(excel_data1)-1
            co1_Fail,co2_Fail,co3_Fail,co4_Fail,co5_Fail,co6_Fail = 0,0,0,0,0,0
            #co1_Pass,co2_Pass,co3_Pass = 0,0,0
            co1_pass_percentage,co2_pass_percentage,co3_pass_percentage,co4_pass_percentage,co5_pass_percentage,co6_pass_percentage = 0,0,0,0,0,0
            co1_PO,co2_PO,co3_PO,co4_PO,co5_PO,co6_PO = 0,0,0,0,0,0
            assign1,assign2 = [],[]
            for i in range(1,len(excel_data1)):
                if excel_data1[i][2]=='AB' or excel_data1[i][2]=='ab' or excel_data1[i][2]=='Ab': excel_data1[i][2]=0
                if excel_data1[i][3]=='AB' or excel_data1[i][3]=='ab' or excel_data1[i][3]=='Ab': excel_data1[i][3]=0
                if excel_data1[i][4]=='AB' or excel_data1[i][4]=='ab' or excel_data1[i][4]=='Ab': excel_data1[i][4]=0
                excel_data1[i].append(round((int(excel_data1[i][2])+int(excel_data1[i][3])+int(excel_data1[i][4]))/3,2))
                assign1.append(type(round((int(excel_data1[i][2])+int(excel_data1[i][3])+int(excel_data1[i][4]))/3,2)))
                a = (int(excel_data1[i][2])/5)*100
                b = (int(excel_data1[i][3])/5)*100
                c = (int(excel_data1[i][4])/5)*100
                
                Gr_a = grade(a)
                Gr_b = grade(b)
                Gr_c = grade(c)
                
                if Gr_a == 'F': co1_Fail += 1
                if Gr_b == 'F': co2_Fail += 1
                if Gr_c == 'F': co3_Fail += 1
                
                #appending Grades
                excel_data1[i].append(Gr_a)
                excel_data1[i].append(Gr_b)
                excel_data1[i].append(Gr_c)
                

                #appending assignment 2 total
                if excel_data1[i][5]=='AB' or excel_data1[i][5]=='ab' or excel_data1[i][5]=='Ab': excel_data1[i][5]=0
                if excel_data1[i][6]=='AB' or excel_data1[i][6]=='ab' or excel_data1[i][6]=='Ab': excel_data1[i][6]=0
                if excel_data1[i][7]=='AB' or excel_data1[i][7]=='ab' or excel_data1[i][7]=='Ab': excel_data1[i][7]=0
                excel_data1[i].append(round((int(excel_data1[i][5])+int(excel_data1[i][6])+int(excel_data1[i][7]))/3,2))
                assign2.append(round((int(excel_data1[i][5])+int(excel_data1[i][6])+int(excel_data1[i][7]))/3,2))
                d = (int(excel_data1[i][5])/5)*100
                e = (int(excel_data1[i][6])/5)*100
                f = (int(excel_data1[i][7])/5)*100

                Gr_d = grade(d)
                Gr_e = grade(e)
                Gr_f = grade(f)

                if Gr_d == 'F': co4_Fail += 1
                if Gr_e == 'F': co5_Fail += 1
                if Gr_f == 'F': co6_Fail += 1

                #appending assignment 2 grades
                excel_data1[i].append(Gr_d)
                excel_data1[i].append(Gr_e)
                excel_data1[i].append(Gr_f)

            #finding pass percentage
            co1_pass_percentage = round(((Total_Students-co1_Fail)/Total_Students)*100)
            co2_pass_percentage = round(((Total_Students-co2_Fail)/Total_Students)*100)
            co3_pass_percentage = round(((Total_Students-co3_Fail)/Total_Students)*100)
            co4_pass_percentage = round(((Total_Students-co4_Fail)/Total_Students)*100)
            co5_pass_percentage = round(((Total_Students-co5_Fail)/Total_Students)*100)
            co6_pass_percentage = round(((Total_Students-co6_Fail)/Total_Students)*100)
            
            #Finding Final PO Values
            co1_PO = mainPO(co1_pass_percentage)
            co2_PO = mainPO(co2_pass_percentage)
            co3_PO = mainPO(co3_pass_percentage)
            co4_PO = mainPO(co4_pass_percentage)
            co5_PO = mainPO(co5_pass_percentage)
            co6_PO = mainPO(co6_pass_percentage)
        
            Main_excel_sheet['Assignment'] = [co1_PO,co2_PO,co3_PO,co4_PO,co5_PO,co6_PO]
            #appending Pass percentage
            l1 = ['']*len(excel_data1[1])
            l1[9] = (co1_pass_percentage)
            l1[10] = (co2_pass_percentage)
            l1[11] = (co3_pass_percentage)
            l1[13] = (co4_pass_percentage)
            l1[14] = (co5_pass_percentage)
            l1[15] = (co6_pass_percentage)
            excel_data1.append(l1)

            #appending Final PO Values
            l1 = ['']*len(excel_data1[1])
            l1[9] = (co1_PO)
            l1[10] = (co2_PO)
            l1[11] = (co3_PO)
            l1[13] = (co4_PO)
            l1[14] = (co5_PO)
            l1[15] = (co6_PO)
            excel_data1.append(l1)


            # for row in worksheet2.iter_rows():
            #     row_data = list()
            #     for cell in row:
            #         row_data.append(str(cell.value))
            #     excel_data2.append(row_data)
            list02 = ['Total Internal MID 1','Internal_Mid1_Grade','Grade(CO1)','Grade(CO2)','Grade(CO3)','Quiz1_Grade','Total Internal MID 2','Internal_Mid2_Grade','Grade(CO4)','Grade(CO5)','Grade(CO6)','Quiz2_Grade']
            excel_data2[0].extend(list02)

            mid1_total,mid1_percentage,mid2_total,mid2_percentage = 0,0,0,0
            co1_mid,co2_mid,co3_mid,co4_mid,co5_mid,co6_mid = 0,0,0,0,0,0
            quiz1,quiz2 = 0,0
            co1_mid_fail,co2_mid_fail,co3_mid_fail,co4_mid_fail,co5_mid_fail,co6_mid_fail = 0,0,0,0,0,0
            quiz1_mid_fail,quiz2_mid_fail = 0,0
            co1_mid_pass_percentage,co2_mid_pass_percentage,co3_mid_pass_percentage,co4_mid_pass_percentage,co5_mid_pass_percentage,co6_mid_pass_percentage = 0,0,0,0,0,0
            co1_mid_PO,co2_mid_PO,co3_mid_PO,co4_mid_PO,co5_mid_PO,co6_mid_PO = 0,0,0,0,0,0
            quiz1_pass_percentage, quiz2_pass_percentage = 0,0
            quiz1_PO,quiz2_PO = 0,0,


            for i in range(1,len(excel_data2)):
                if excel_data2[i][2]=='AB' or excel_data2[i][2]=='ab' or excel_data2[i][2]=='Ab' or excel_data2[i][2]=='': 
                    excel_data2[i][2]=0
                if excel_data2[i][3]=='AB' or excel_data2[i][3]=='ab' or excel_data2[i][3]=='Ab' or excel_data2[i][3]=='': 
                    excel_data2[i][3]=0
                if excel_data2[i][4]=='AB' or excel_data2[i][4]=='ab' or excel_data2[i][4]=='Ab' or excel_data2[i][4]=='': 
                    excel_data2[i][4]=0
                if excel_data2[i][5]=='AB' or excel_data2[i][5]=='ab' or excel_data2[i][5]=='Ab' or excel_data2[i][5]=='': 
                    excel_data2[i][5]=0
                mid1_total = ceil(float(excel_data2[i][2])+float(excel_data2[i][3])+float(excel_data2[i][4]))+int(excel_data2[i][5])+float(excel_data1[i][8])
                excel_data2[i].append(mid1_total)
                mid1_percentage = (mid1_total/25)*100
                excel_data2[i].append(grade(mid1_percentage))
                co1_mid = (float(excel_data2[i][2])/4)*100
                co2_mid = (float(excel_data2[i][3])/4)*100
                co3_mid = (float(excel_data2[i][4])/2)*100
                quiz1 = (int(excel_data2[i][5])/10)*100

                Grade_mid_co1 = grade(co1_mid)
                Grade_mid_co2 = grade(co2_mid)
                Grade_mid_co3 = grade(co3_mid)
                Grade_quiz1 = grade(quiz1)

                if Grade_mid_co1 == "F" : co1_mid_fail += 1
                if Grade_mid_co2 == "F" : co2_mid_fail += 1
                if Grade_mid_co3 == "F" : co3_mid_fail += 1
                if Grade_quiz1 == "F" : quiz1_mid_fail += 1
                
                excel_data2[i].append(Grade_mid_co1)
                excel_data2[i].append(Grade_mid_co2)
                excel_data2[i].append(Grade_mid_co3)
                excel_data2[i].append(Grade_quiz1)


                if excel_data2[i][6]=='AB' or excel_data2[i][6]=='ab' or excel_data2[i][6]=='Ab' or  excel_data2[i][6]=='': excel_data2[i][6]=0
                if excel_data2[i][7]=='AB' or excel_data2[i][7]=='ab' or excel_data2[i][7]=='Ab' or excel_data2[i][7]=='': excel_data2[i][7]=0
                if excel_data2[i][8]=='AB' or excel_data2[i][8]=='ab' or excel_data2[i][8]=='Ab' or excel_data2[i][6]=='': excel_data2[i][8]=0
                if excel_data2[i][9]=='AB' or excel_data2[i][9]=='ab' or excel_data2[i][9]=='Ab' or excel_data2[i][6]=='': excel_data2[i][9]=0
                mid2_total = ceil(float(excel_data2[i][6])+float(excel_data2[i][7])+float(excel_data2[i][8]))+int(excel_data2[i][9])+float((excel_data1[i][12]))
                excel_data2[i].append(mid2_total)
                mid2_percentage = (mid2_total/25)*100
                excel_data2[i].append(grade(mid2_percentage))
                co4_mid = (float(excel_data2[i][6])/2)*100
                co5_mid = (float(excel_data2[i][7])/4)*100
                co6_mid = (float(excel_data2[i][8])/4)*100
                quiz2 = (int(excel_data2[i][9])/10)*100

                Grade_mid_co4 = grade(co4_mid)
                Grade_mid_co5 = grade(co5_mid)
                Grade_mid_co6 = grade(co6_mid)
                Grade_quiz2 = grade(quiz2)

                if Grade_mid_co4 == "F" : co4_mid_fail += 1
                if Grade_mid_co5 == "F" : co5_mid_fail += 1
                if Grade_mid_co6 == "F" : co6_mid_fail += 1
                if Grade_quiz2 == "F" : quiz2_mid_fail += 1
                
                excel_data2[i].append(Grade_mid_co4)
                excel_data2[i].append(Grade_mid_co5)
                excel_data2[i].append(Grade_mid_co6)
                excel_data2[i].append(Grade_quiz2)


            #calculating pass percentage for mids and quiz
            co1_mid_pass_percentage = round(((Total_Students-co1_mid_fail)/Total_Students)*100)
            co2_mid_pass_percentage = round(((Total_Students-co2_mid_fail)/Total_Students)*100)
            co3_mid_pass_percentage = round(((Total_Students-co3_mid_fail)/Total_Students)*100)
            quiz1_pass_percentage = round(((Total_Students-quiz1_mid_fail)/Total_Students)*100)
            co4_mid_pass_percentage = round(((Total_Students-co4_mid_fail)/Total_Students)*100)
            co5_mid_pass_percentage = round(((Total_Students-co5_mid_fail)/Total_Students)*100)
            co6_mid_pass_percentage = round(((Total_Students-co6_mid_fail)/Total_Students)*100)
            quiz2_pass_percentage = round(((Total_Students-quiz2_mid_fail)/Total_Students)*100)

            co1_mid_PO = mainPO(co1_mid_pass_percentage)
            co2_mid_PO = mainPO(co2_mid_pass_percentage)
            co3_mid_PO = mainPO(co3_mid_pass_percentage)
            quiz1_PO = mainPO(quiz1_pass_percentage)
            co4_mid_PO = mainPO(co4_mid_pass_percentage)
            co5_mid_PO = mainPO(co5_mid_pass_percentage)
            co6_mid_PO = mainPO(co6_mid_pass_percentage)
            quiz2_PO = mainPO(quiz2_pass_percentage)

            l1 = ['']*len(excel_data2[1])
            l1[12] = (co1_mid_pass_percentage)
            l1[13] = (co2_mid_pass_percentage)
            l1[14] = (co3_mid_pass_percentage)
            l1[15] = (quiz1_pass_percentage)
            l1[18] = (co4_mid_pass_percentage)
            l1[19] = (co5_mid_pass_percentage)
            l1[20] = (co6_mid_pass_percentage)
            l1[21] = (quiz2_pass_percentage)
            excel_data2.append(l1)

            l1 = ['']*len(excel_data2[1])
            l1[12] = (co1_mid_PO)
            l1[13] = (co2_mid_PO)
            l1[14] = (co3_mid_PO)
            l1[15] = (quiz1_PO)
            l1[18] = (co4_mid_PO)
            l1[19] = (co5_mid_PO)
            l1[20] = (co6_mid_PO)
            l1[21] = (quiz2_PO)
            excel_data2.append(l1)

            Main_excel_sheet['Mid'] = [co1_mid_PO,co2_mid_PO,co3_mid_PO,co4_mid_PO,co5_mid_PO,co6_mid_PO]
            Main_excel_sheet['Quiz'] = [quiz1_PO,quiz2_PO]


            #Sheet3

            # for row in worksheet3.iter_rows():
            #     row_data = list()
            #     for cell in row:
            #         row_data.append(str(cell.value))
            #     excel_data3.append(row_data)

            list03 = ['Grade point','Achieved Target Level']
            excel_data3[0].extend(list03)
            sem_fail_count = 0
            sem_pass_count = 0
            sem_pass_percenatge = 0
            sem_PO = 0
            sem_absent = 0
            for i in range(1,len(excel_data3)):
                if excel_data3[i][2] == 'F':
                    excel_data3[i].append(gradepoint('F'))
                    excel_data3[i].append('NO')
                    sem_fail_count += 1
                elif excel_data3[i][2] == 'AB':
                    excel_data3[i].append(gradepoint('AB'))
                    excel_data3[i].append('NO')
                    sem_absent += 1
                else:
                    excel_data3[i].append(gradepoint(excel_data3[i][2]))
                    excel_data3[i].append('YES')
                    sem_pass_count += 1
                
            sem_pass_percenatge = round((sem_pass_count/(Total_Students-sem_absent))*100)
            sem_PO = mainPO(sem_pass_percenatge)

            l1 = ['']*len(excel_data3[1])
            l1[3] = ('Sem pass Percentage = ')
            l1[4] = (sem_pass_percenatge)
            excel_data3.append(l1)

            l1 = ['']*len(excel_data3[1])
            l1[3] = ('Sem PO = ')
            l1[4] = (sem_PO)
            excel_data3.append(l1)

            Main_excel_sheet['Sem'] = [sem_PO]


            #Sheet4

            for row in worksheet4.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data4.append(row_data)
            
            list04 = ['Assignment','Quiz','Descriptive','CO wise attainment','SEE','Direct CO attainment','Overall CO attainment']
            excel_data4[0].extend(list04)


            assign = Main_excel_sheet['Assignment']
            quiz = Main_excel_sheet['Quiz']
            mid = Main_excel_sheet['Mid']
            sem = Main_excel_sheet['Sem']

            co_wise_attainment = []
            direct_co_attainment = []
            overall_co_attainment = []


            for i in range(1,7):
                temp1 = assign[i-1]
                excel_data4[i].append(assign[i-1])
                if i<4: 
                    temp2 = quiz[0]
                    excel_data4[i].append(quiz[0])
                else: 
                    temp2 = quiz[1]
                    excel_data4[i].append(quiz[1])
                temp3 = mid[i-1]
                excel_data4[i].append(mid[i-1])
                co_wise = round((temp1+temp2+temp3)/3)
                co_wise_attainment.append(co_wise)
                excel_data4[i].append(co_wise)
                excel_data4[i].append(sem[0])
                direct_co = 0.5*co_wise+0.5*sem[0]
                excel_data4[i].append(direct_co)
                direct_co_attainment.append(direct_co)
                overall = round(0.80*direct_co + 0.20*float(excel_data4[i][1]),2)
                excel_data4[i].append(overall)
                overall_co_attainment.append(overall)
            average_co = round(sum(overall_co_attainment)/len(overall_co_attainment),3)
            excel_data4.append(['','','','','','','','Average CO Attainment',average_co])


            #Sheet 5

            for row in worksheet5.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data5.append(row_data) 

            excel_data6 += excel_data5          

            PO1,PO2,PO3,PO4,PO5,PO6,PO7,PO8,PO9,PO10,PO11,PO12,PSO1,PSO2,PSO3 = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
            PO1_II,PO2_II,PO3_II,PO4_II,PO5_II,PO6_II,PO7_II,PO8_II,PO9_II,PO10_II,PO11_II,PO12_II,PSO1_II,PSO2_II,PSO3_II = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
            PO_Average,PSO_Average = [],[]
            POII_Average,PSOII_Average = [],[]
            for i in range(2,len(excel_data5)):
                if none_or_not(excel_data5[i][2]):
                    PO1.append(int(excel_data5[i][2]))
                    PO1_II.append(round((int(excel_data5[i][2])*average_co)/3,2))
                if none_or_not(excel_data5[i][3]):
                    PO2.append(int(excel_data5[i][3]))
                    PO2_II.append(round((int(excel_data5[i][3])*average_co)/3,2))
                if none_or_not(excel_data5[i][4]):
                    PO3.append(int(excel_data5[i][4]))
                    PO3_II.append(round((int(excel_data5[i][4])*average_co)/3,2))
                if none_or_not(excel_data5[i][5]):
                    PO4.append(int(excel_data5[i][5]))
                    PO4_II.append(round((int(excel_data5[i][5])*average_co)/3,2))
                if none_or_not(excel_data5[i][6]):
                    PO5.append(int(excel_data5[i][6]))
                    PO5_II.append(round((int(excel_data5[i][6])*average_co)/3,2))
                if none_or_not(excel_data5[i][7]):
                    PO6.append(int(excel_data5[i][7]))
                    PO6_II.append(round((int(excel_data5[i][7])*average_co)/3,2))
                if none_or_not(excel_data5[i][8]):
                    PO7.append(int(excel_data5[i][8]))
                    PO7_II.append(round((int(excel_data5[i][8])*average_co)/3,2))
                if none_or_not(excel_data5[i][9]):
                    PO8.append(int(excel_data5[i][9]))
                    PO8_II.append(round((int(excel_data5[i][9])*average_co)/3,2))
                if none_or_not(excel_data5[i][10]):
                    PO9.append(int(excel_data5[i][10]))
                    PO9_II.append(round((int(excel_data5[i][10])*average_co)/3,2))
                if none_or_not(excel_data5[i][11]):
                    PO10.append(int(excel_data5[i][11]))
                    PO10_II.append(round((int(excel_data5[i][11])*average_co)/3,2))
                if none_or_not(excel_data5[i][12]):
                    PO11.append(int(excel_data5[i][12]))
                    PO11_II.append(round((int(excel_data5[i][12])*average_co)/3,2))
                if none_or_not(excel_data5[i][13]):
                    PO12.append(int(excel_data5[i][13]))
                    PO12_II.append(round((int(excel_data5[i][13])*average_co)/3,2))
                if none_or_not(excel_data5[i][14]):
                    PSO1.append(int(excel_data5[i][14]))
                    PSO1_II.append(round((int(excel_data5[i][14])*average_co)/3,2))
                if none_or_not(excel_data5[i][15]):
                    PSO2.append(int(excel_data5[i][15]))
                    PSO2_II.append(round((int(excel_data5[i][15])*average_co)/3,2))
                if none_or_not(excel_data5[i][16]):
                    PSO3.append(int(excel_data5[i][16]))
                    PSO3_II.append(round((int(excel_data5[i][16])*average_co)/3,2))
            

            if len(PO1): 
                PO_Average.append(round(sum(PO1)/len(PO1),2))
                POII_Average.append(round(sum(PO1_II)/len(PO1_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)
            
            if len(PO2): 
                PO_Average.append(round(sum(PO2)/len(PO2),2)) 
                POII_Average.append(round(sum(PO2_II)/len(PO2_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO3): 
                PO_Average.append(round(sum(PO3)/len(PO3),2)) 
                POII_Average.append(round(sum(PO3_II)/len(PO3_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO4): 
                PO_Average.append(round(sum(PO4)/len(PO4),2))
                POII_Average.append(round(sum(PO4_II)/len(PO4_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO5): 
                PO_Average.append(round(sum(PO5)/len(PO5),2))
                POII_Average.append(round(sum(PO5_II)/len(PO5_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO6): 
                PO_Average.append(round(sum(PO6)/len(PO6),2)) 
                POII_Average.append(round(sum(PO6_II)/len(PO6_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO7): 
                PO_Average.append(round(sum(PO7)/len(PO7),2)) 
                POII_Average.append(round(sum(PO7_II)/len(PO7_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO8): 
                PO_Average.append(round(sum(PO8)/len(PO8),2)) 
                POII_Average.append(round(sum(PO8_II)/len(PO8_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO9): 
                PO_Average.append(round(sum(PO9)/len(PO9),2))
                POII_Average.append(round(sum(PO9_II)/len(PO9_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO10): 
                PO_Average.append(round(sum(PO10)/len(PO10),2))
                POII_Average.append(round(sum(PO10_II)/len(PO10_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO11): 
                PO_Average.append(round(sum(PO11)/len(PO11),2))
                POII_Average.append(round(sum(PO11_II)/len(PO11_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO12): 
                PO_Average.append(round(sum(PO12)/len(PO12),2))
                POII_Average.append(round(sum(PO12_II)/len(PO12_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PSO1): 
                PSO_Average.append(round(sum(PSO1)/len(PSO1),2)) 
                PSOII_Average.append(round(sum(PSO1_II)/len(PSO1_II),2))
            else:
                PSO_Average.append(0)
                PSOII_Average.append(0)

            if len(PSO2): 
                PSO_Average.append(round(sum(PSO2)/len(PSO2),2)) 
                PSOII_Average.append(round(sum(PSO2_II)/len(PSO2_II),2))
            else:
                PSO_Average.append(0)
                PSOII_Average.append(0)

            if len(PSO3): 
                PSO_Average.append(round(sum(PSO3)/len(PSO3),2))
                PSOII_Average.append(round(sum(PSO3_II)/len(PSO3_II),2)) 
            else:
                PSO_Average.append(0)
                PSOII_Average.append(0)

            # if len(PO1): PSO_Average.append(round(sum(PO1)/len(PO1),2)) 
            # else:PSO_Average.append(0)

            # if len(PO1): PSO_Average.append(round(sum(PO1)/len(PO1),2)) 
            # else:PSO_Average.append(0)


            
            # for i in range(2,len())
            excel_data5.append(['','']+PO_Average+PSO_Average)
            excel_data5.append(['','']+POII_Average+PSOII_Average)
            denominator_PO,denominator_PSO = 0,0
            for i in POII_Average:
                if i!=0: denominator_PO+=1
            for i in PSOII_Average:
                if i!=0: denominator_PSO+=1
            Overall_PO_Attainment = round(sum(POII_Average)/denominator_PO,2)
            Overall_PSO_Attainment = round(sum(PSOII_Average)/denominator_PSO,2)
            excel_data5.append([Overall_PO_Attainment,Overall_PSO_Attainment])

            for i in range(len(excel_data5)):
                for j in range(len(excel_data5[i])):
                    if excel_data5[i][j]=='None' or excel_data5[i][j]==0:
                        excel_data5[i][j] = ''


            #Sheet6
            # for row in worksheet6.iter_rows():
            #     row_data = list()
            #     for cell in row:
            #         row_data.append(str(cell.value))
            #     excel_data6.append(row_data)     
            

            S_PO1,S_PO2,S_PO3,S_PO4,S_PO5,S_PO6,S_PO7,S_PO8,S_PO9,S_PO10,S_PO11,S_PO12,S_PSO1,S_PSO2,S_PSO3 = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
            # S_PO1_II,S_PO2_II,S_PO3_II,S_PO4_II,S_PO5_II,S_PO6_II,S_PO7_II,S_PO8_II,S_PO9_II,S_PO10_II,S_PO11_II,S_PO12_II,S_PSO1_II,S_PSO2_II,S_PSO3_II = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
            S_PO_Average,S_PSO_Average = [],[]
            # S_POII_Average,S_PSOII_Average = [],[]
            for i in range(2,len(excel_data6)):
                if none_or_not(excel_data6[i][2]):
                    S_PO1.append(int(excel_data6[i][2]))
                if none_or_not(excel_data6[i][3]):
                    S_PO2.append(int(excel_data6[i][3]))
                if none_or_not(excel_data6[i][4]):
                    S_PO3.append(int(excel_data6[i][4]))
                if none_or_not(excel_data6[i][5]):
                    S_PO4.append(int(excel_data6[i][5]))
                if none_or_not(excel_data6[i][6]):
                    S_PO5.append(int(excel_data6[i][6]))
                if none_or_not(excel_data6[i][7]):
                    S_PO6.append(int(excel_data6[i][7]))
                if none_or_not(excel_data6[i][8]):
                    S_PO7.append(int(excel_data6[i][8]))
                if none_or_not(excel_data6[i][9]):
                    S_PO8.append(int(excel_data6[i][9]))
                if none_or_not(excel_data6[i][10]):
                    S_PO9.append(int(excel_data6[i][10]))
                if none_or_not(excel_data6[i][11]):
                    S_PO10.append(int(excel_data6[i][11]))
                if none_or_not(excel_data6[i][12]):
                    S_PO11.append(int(excel_data6[i][12]))
                if none_or_not(excel_data6[i][13]):
                    S_PO12.append(int(excel_data6[i][13]))
                if none_or_not(excel_data6[i][14]):
                    S_PSO1.append(int(excel_data6[i][14]))
                if none_or_not(excel_data6[i][15]):
                    S_PSO2.append(int(excel_data6[i][15]))
                if none_or_not(excel_data6[i][16]):
                    S_PSO3.append(int(excel_data6[i][16]))
                
            

            if len(S_PO1): 
                S_PO_Average.append(round(sum(S_PO1)/len(S_PO1),2))
            else:
                S_PO_Average.append(0)
            
            if len(S_PO2): 
                S_PO_Average.append(round(sum(S_PO2)/len(S_PO2),2)) 
            else:
                S_PO_Average.append(0)

            if len(S_PO3): 
                S_PO_Average.append(round(sum(S_PO3)/len(S_PO3),2)) 
            else:
                S_PO_Average.append(0)

            if len(S_PO4): 
                S_PO_Average.append(round(sum(S_PO4)/len(S_PO4),2))
            else:
                S_PO_Average.append(0)

            if len(S_PO5): 
                S_PO_Average.append(round(sum(S_PO5)/len(S_PO5),2))
            else:
                S_PO_Average.append(0)

            if len(S_PO6): 
                S_PO_Average.append(round(sum(S_PO6)/len(S_PO6),2)) 
            else:
                S_PO_Average.append(0)

            if len(S_PO7): 
                S_PO_Average.append(round(sum(S_PO7)/len(S_PO7),2)) 
            else:
                S_PO_Average.append(0)

            if len(S_PO8): 
                S_PO_Average.append(round(sum(S_PO8)/len(S_PO8),2)) 
            else:
                S_PO_Average.append(0)

            if len(S_PO9): 
                S_PO_Average.append(round(sum(S_PO9)/len(S_PO9),2))
            else:
                S_PO_Average.append(0)

            if len(S_PO10): 
                S_PO_Average.append(round(sum(S_PO10)/len(S_PO10),2))
            else:
                S_PO_Average.append(0)

            if len(S_PO11): 
                S_PO_Average.append(round(sum(S_PO11)/len(S_PO11),2))
            else:
                S_PO_Average.append(0)

            if len(S_PO12): 
                S_PO_Average.append(round(sum(S_PO12)/len(S_PO12),2))
            else:
                S_PO_Average.append(0)

            if len(S_PSO1): 
                S_PSO_Average.append(round(sum(S_PSO1)/len(S_PSO1),2)) 
            else:
                S_PSO_Average.append(0)

            if len(S_PSO2): 
                S_PSO_Average.append(round(sum(S_PSO2)/len(S_PSO2),2)) 
            else:
                S_PSO_Average.append(0)

            if len(S_PSO3): 
                S_PSO_Average.append(round(sum(S_PSO3)/len(S_PSO3),2))
            else:
                S_PSO_Average.append(0)

        


            
            # for i in range(2,len())
            excel_data6.append(['']+S_PO_Average+S_PSO_Average)

            for i in range(len(excel_data6)):
                for j in range(len(excel_data6[i])):
                    if excel_data6[i][j]=='None' or excel_data6[i][j]==0:
                        excel_data6[i][j] = ''
            


            return render(request,'po_output.html', {"depart":depart[Branch],"assign2":assign2,"final":average_co,"over":overall_co_attainment,"direct":direct_co_attainment,"internal":co_wise_attainment,"main":Main_excel_sheet,"excel_data1":excel_data1,"excel_data2":excel_data2,"excel_data3":excel_data3,"excel_data4":excel_data4,"excel_data5":excel_data5,"Details":details,"excel_data6":excel_data6})
            #return render(request,'po_output.html',{"excel_data1":excel_data1,"excel_data2":excel_data2,"excel_data3":excel_data3})

        elif Regulation=='R20':
                #print(worksheet)
            Main_excel_sheet = {}
            excel_data = list()
            excel_data1 = list()
            excel_data2 = list()
            excel_data3 = list()
            excel_data4 = list()
            excel_data5 = list()
            excel_data6 = list()
            count,i = 0,0
            nums = ['0','1','2','3','4','5','6','7','8','9']
            # iterating over the rows and
            # getting value from each cell in row
            for row in worksheet1.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data.append(row_data)

            for row in excel_data:
                excel_data1.append(row[0:8])
                excel_data2.append(row[0:2])
                excel_data2[i].extend(row[8:16])
                excel_data3.append(row[0:2])
                excel_data3[i].extend(row[-1])
                i+=1
            
            list01 = ['Total Assignment - 1',"CO1 (Grade)","CO2 (Grade)","CO3 (Grade)",'Total Assignment - 2',"CO4 (Grade)","CO5 (Grade)","CO6 (Grade)"]
            excel_data1[0].extend(list01)
        

            Total_Students = len(excel_data1)-1
            co1_Fail,co2_Fail,co3_Fail,co4_Fail,co5_Fail,co6_Fail = 0,0,0,0,0,0
            #co1_Pass,co2_Pass,co3_Pass = 0,0,0
            co1_pass_percentage,co2_pass_percentage,co3_pass_percentage,co4_pass_percentage,co5_pass_percentage,co6_pass_percentage = 0,0,0,0,0,0
            co1_PO,co2_PO,co3_PO,co4_PO,co5_PO,co6_PO = 0,0,0,0,0,0
            assign1,assign2 = [],[]
            for i in range(1,len(excel_data1)):
                if excel_data1[i][2]=='AB' or excel_data1[i][2]=='ab' or excel_data1[i][2]=='Ab': excel_data1[i][2]=0
                if excel_data1[i][3]=='AB' or excel_data1[i][3]=='ab' or excel_data1[i][3]=='Ab': excel_data1[i][3]=0
                if excel_data1[i][4]=='AB' or excel_data1[i][4]=='ab' or excel_data1[i][4]=='Ab': excel_data1[i][4]=0
                excel_data1[i].append(round((int(excel_data1[i][2])+int(excel_data1[i][3])+int(excel_data1[i][4]))/3,2))
                assign1.append(type(round((int(excel_data1[i][2])+int(excel_data1[i][3])+int(excel_data1[i][4]))/3,2)))
                a = (int(excel_data1[i][2])/5)*100
                b = (int(excel_data1[i][3])/5)*100
                c = (int(excel_data1[i][4])/5)*100
                
                Gr_a = gradeR20(a)
                Gr_b = gradeR20(b)
                Gr_c = gradeR20(c)
                
                if Gr_a == 'F': co1_Fail += 1
                if Gr_b == 'F': co2_Fail += 1
                if Gr_c == 'F': co3_Fail += 1
                
                #appending Grades
                excel_data1[i].append(Gr_a)
                excel_data1[i].append(Gr_b)
                excel_data1[i].append(Gr_c)
                

                #appending assignment 2 total
                if excel_data1[i][5]=='AB' or excel_data1[i][5]=='ab' or excel_data1[i][5]=='Ab': excel_data1[i][5]=0
                if excel_data1[i][6]=='AB' or excel_data1[i][6]=='ab' or excel_data1[i][6]=='Ab': excel_data1[i][6]=0
                if excel_data1[i][7]=='AB' or excel_data1[i][7]=='ab' or excel_data1[i][7]=='Ab': excel_data1[i][7]=0
                excel_data1[i].append(round((int(excel_data1[i][5])+int(excel_data1[i][6])+int(excel_data1[i][7]))/3,2))
                assign2.append(round((int(excel_data1[i][5])+int(excel_data1[i][6])+int(excel_data1[i][7]))/3,2))
                d = (int(excel_data1[i][5])/5)*100
                e = (int(excel_data1[i][6])/5)*100
                f = (int(excel_data1[i][7])/5)*100

                Gr_d = gradeR20(d)
                Gr_e = gradeR20(e)
                Gr_f = gradeR20(f)

                if Gr_d == 'F': co4_Fail += 1
                if Gr_e == 'F': co5_Fail += 1
                if Gr_f == 'F': co6_Fail += 1

                #appending assignment 2 grades
                excel_data1[i].append(Gr_d)
                excel_data1[i].append(Gr_e)
                excel_data1[i].append(Gr_f)

            #finding pass percentage
            co1_pass_percentage = round(((Total_Students-co1_Fail)/Total_Students)*100)
            co2_pass_percentage = round(((Total_Students-co2_Fail)/Total_Students)*100)
            co3_pass_percentage = round(((Total_Students-co3_Fail)/Total_Students)*100)
            co4_pass_percentage = round(((Total_Students-co4_Fail)/Total_Students)*100)
            co5_pass_percentage = round(((Total_Students-co5_Fail)/Total_Students)*100)
            co6_pass_percentage = round(((Total_Students-co6_Fail)/Total_Students)*100)
            
            #Finding Final PO Values
            co1_PO = mainPO(co1_pass_percentage)
            co2_PO = mainPO(co2_pass_percentage)
            co3_PO = mainPO(co3_pass_percentage)
            co4_PO = mainPO(co4_pass_percentage)
            co5_PO = mainPO(co5_pass_percentage)
            co6_PO = mainPO(co6_pass_percentage)
        
            Main_excel_sheet['Assignment'] = [co1_PO,co2_PO,co3_PO,co4_PO,co5_PO,co6_PO]
            #appending Pass percentage
            l1 = ['']*len(excel_data1[1])
            l1[9] = (co1_pass_percentage)
            l1[10] = (co2_pass_percentage)
            l1[11] = (co3_pass_percentage)
            l1[13] = (co4_pass_percentage)
            l1[14] = (co5_pass_percentage)
            l1[15] = (co6_pass_percentage)
            excel_data1.append(l1)

            #appending Final PO Values
            l1 = ['']*len(excel_data1[1])
            l1[9] = (co1_PO)
            l1[10] = (co2_PO)
            l1[11] = (co3_PO)
            l1[13] = (co4_PO)
            l1[14] = (co5_PO)
            l1[15] = (co6_PO)
            excel_data1.append(l1)


            # for row in worksheet2.iter_rows():
            #     row_data = list()
            #     for cell in row:
            #         row_data.append(str(cell.value))
            #     excel_data2.append(row_data)
            list02 = ['Total Internal MID 1','Internal_Mid1_Grade','Grade(CO1)','Grade(CO2)','Grade(CO3)','Quiz1_Grade','Total Internal MID 2','Internal_Mid2_Grade','Grade(CO4)','Grade(CO5)','Grade(CO6)','Quiz2_Grade']
            excel_data2[0].extend(list02)

            mid1_total,mid1_percentage,mid2_total,mid2_percentage = 0,0,0,0
            co1_mid,co2_mid,co3_mid,co4_mid,co5_mid,co6_mid = 0,0,0,0,0,0
            quiz1,quiz2 = 0,0
            co1_mid_fail,co2_mid_fail,co3_mid_fail,co4_mid_fail,co5_mid_fail,co6_mid_fail = 0,0,0,0,0,0
            quiz1_mid_fail,quiz2_mid_fail = 0,0
            co1_mid_pass_percentage,co2_mid_pass_percentage,co3_mid_pass_percentage,co4_mid_pass_percentage,co5_mid_pass_percentage,co6_mid_pass_percentage = 0,0,0,0,0,0
            co1_mid_PO,co2_mid_PO,co3_mid_PO,co4_mid_PO,co5_mid_PO,co6_mid_PO = 0,0,0,0,0,0
            quiz1_pass_percentage, quiz2_pass_percentage = 0,0
            quiz1_PO,quiz2_PO = 0,0,


            for i in range(1,len(excel_data2)):
                if excel_data2[i][2]=='AB' or excel_data2[i][2]=='ab' or excel_data2[i][2]=='Ab' or excel_data2[i][2]=='': 
                    excel_data2[i][2]=0
                if excel_data2[i][3]=='AB' or excel_data2[i][3]=='ab' or excel_data2[i][3]=='Ab' or excel_data2[i][3]=='': 
                    excel_data2[i][3]=0
                if excel_data2[i][4]=='AB' or excel_data2[i][4]=='ab' or excel_data2[i][4]=='Ab' or excel_data2[i][4]=='': 
                    excel_data2[i][4]=0
                if excel_data2[i][5]=='AB' or excel_data2[i][5]=='ab' or excel_data2[i][5]=='Ab' or excel_data2[i][5]=='': 
                    excel_data2[i][5]=0
                mid1_total = ceil(float(excel_data2[i][2])+float(excel_data2[i][3])+float(excel_data2[i][4]))+int(excel_data2[i][5])+float(excel_data1[i][8])
                excel_data2[i].append(mid1_total)
                mid1_percentage = (mid1_total/25)*100
                excel_data2[i].append(gradeR20(mid1_percentage))
                co1_mid = (float(excel_data2[i][2])/5)*100
                co2_mid = (float(excel_data2[i][3])/5)*100
                co3_mid = (float(excel_data2[i][4])/5)*100
                quiz1 = (int(excel_data2[i][5])/10)*100

                Grade_mid_co1 = gradeR20(co1_mid)
                Grade_mid_co2 = gradeR20(co2_mid)
                Grade_mid_co3 = gradeR20(co3_mid)
                Grade_quiz1 = gradeR20(quiz1)

                if Grade_mid_co1 == "F" : co1_mid_fail += 1
                if Grade_mid_co2 == "F" : co2_mid_fail += 1
                if Grade_mid_co3 == "F" : co3_mid_fail += 1
                if Grade_quiz1 == "F" : quiz1_mid_fail += 1
                
                excel_data2[i].append(Grade_mid_co1)
                excel_data2[i].append(Grade_mid_co2)
                excel_data2[i].append(Grade_mid_co3)
                excel_data2[i].append(Grade_quiz1)


                if excel_data2[i][6]=='AB' or excel_data2[i][6]=='ab' or excel_data2[i][6]=='Ab' or  excel_data2[i][6]=='': excel_data2[i][6]=0
                if excel_data2[i][7]=='AB' or excel_data2[i][7]=='ab' or excel_data2[i][7]=='Ab' or excel_data2[i][7]=='': excel_data2[i][7]=0
                if excel_data2[i][8]=='AB' or excel_data2[i][8]=='ab' or excel_data2[i][8]=='Ab' or excel_data2[i][6]=='': excel_data2[i][8]=0
                if excel_data2[i][9]=='AB' or excel_data2[i][9]=='ab' or excel_data2[i][9]=='Ab' or excel_data2[i][6]=='': excel_data2[i][9]=0
                mid2_total = ceil(float(excel_data2[i][6])+float(excel_data2[i][7])+float(excel_data2[i][8]))+int(excel_data2[i][9])+float((excel_data1[i][12]))
                excel_data2[i].append(mid2_total)
                mid2_percentage = (mid2_total/25)*100
                excel_data2[i].append(gradeR20(mid2_percentage))
                co4_mid = (float(excel_data2[i][6])/5)*100
                co5_mid = (float(excel_data2[i][7])/5)*100
                co6_mid = (float(excel_data2[i][8])/5)*100
                quiz2 = (int(excel_data2[i][9])/10)*100

                Grade_mid_co4 = gradeR20(co4_mid)
                Grade_mid_co5 = gradeR20(co5_mid)
                Grade_mid_co6 = gradeR20(co6_mid)
                Grade_quiz2 = gradeR20(quiz2)

                if Grade_mid_co4 == "F" : co4_mid_fail += 1
                if Grade_mid_co5 == "F" : co5_mid_fail += 1
                if Grade_mid_co6 == "F" : co6_mid_fail += 1
                if Grade_quiz2 == "F" : quiz2_mid_fail += 1
                
                excel_data2[i].append(Grade_mid_co4)
                excel_data2[i].append(Grade_mid_co5)
                excel_data2[i].append(Grade_mid_co6)
                excel_data2[i].append(Grade_quiz2)


            #calculating pass percentage for mids and quiz
            co1_mid_pass_percentage = round(((Total_Students-co1_mid_fail)/Total_Students)*100)
            co2_mid_pass_percentage = round(((Total_Students-co2_mid_fail)/Total_Students)*100)
            co3_mid_pass_percentage = round(((Total_Students-co3_mid_fail)/Total_Students)*100)
            quiz1_pass_percentage = round(((Total_Students-quiz1_mid_fail)/Total_Students)*100)
            co4_mid_pass_percentage = round(((Total_Students-co4_mid_fail)/Total_Students)*100)
            co5_mid_pass_percentage = round(((Total_Students-co5_mid_fail)/Total_Students)*100)
            co6_mid_pass_percentage = round(((Total_Students-co6_mid_fail)/Total_Students)*100)
            quiz2_pass_percentage = round(((Total_Students-quiz2_mid_fail)/Total_Students)*100)

            co1_mid_PO = mainPO(co1_mid_pass_percentage)
            co2_mid_PO = mainPO(co2_mid_pass_percentage)
            co3_mid_PO = mainPO(co3_mid_pass_percentage)
            quiz1_PO = mainPO(quiz1_pass_percentage)
            co4_mid_PO = mainPO(co4_mid_pass_percentage)
            co5_mid_PO = mainPO(co5_mid_pass_percentage)
            co6_mid_PO = mainPO(co6_mid_pass_percentage)
            quiz2_PO = mainPO(quiz2_pass_percentage)

            l1 = ['']*len(excel_data2[1])
            l1[12] = (co1_mid_pass_percentage)
            l1[13] = (co2_mid_pass_percentage)
            l1[14] = (co3_mid_pass_percentage)
            l1[15] = (quiz1_pass_percentage)
            l1[18] = (co4_mid_pass_percentage)
            l1[19] = (co5_mid_pass_percentage)
            l1[20] = (co6_mid_pass_percentage)
            l1[21] = (quiz2_pass_percentage)
            excel_data2.append(l1)

            l1 = ['']*len(excel_data2[1])
            l1[12] = (co1_mid_PO)
            l1[13] = (co2_mid_PO)
            l1[14] = (co3_mid_PO)
            l1[15] = (quiz1_PO)
            l1[18] = (co4_mid_PO)
            l1[19] = (co5_mid_PO)
            l1[20] = (co6_mid_PO)
            l1[21] = (quiz2_PO)
            excel_data2.append(l1)

            Main_excel_sheet['Mid'] = [co1_mid_PO,co2_mid_PO,co3_mid_PO,co4_mid_PO,co5_mid_PO,co6_mid_PO]
            Main_excel_sheet['Quiz'] = [quiz1_PO,quiz2_PO]


            #Sheet3

            # for row in worksheet3.iter_rows():
            #     row_data = list()
            #     for cell in row:
            #         row_data.append(str(cell.value))
            #     excel_data3.append(row_data)

            list03 = ['Grade point','Achieved Target Level']
            excel_data3[0].extend(list03)
            sem_fail_count = 0
            sem_pass_count = 0
            sem_pass_percenatge = 0
            sem_PO = 0
            sem_absent = 0
            for i in range(1,len(excel_data3)):
                if excel_data3[i][2] == 'F':
                    excel_data3[i].append(gradepointR20('F'))
                    excel_data3[i].append('NO')
                    sem_fail_count += 1
                elif excel_data3[i][2] == 'AB':
                    excel_data3[i].append(gradepointR20('AB'))
                    excel_data3[i].append('NO')
                    sem_absent += 1
                else:
                    excel_data3[i].append(gradepointR20(excel_data3[i][2]))
                    excel_data3[i].append('YES')
                    sem_pass_count += 1
                
            sem_pass_percenatge = round((sem_pass_count/(Total_Students-sem_absent))*100)
            sem_PO = mainPO(sem_pass_percenatge)

            l1 = ['']*len(excel_data3[1])
            l1[3] = ('Sem pass Percentage = ')
            l1[4] = (sem_pass_percenatge)
            excel_data3.append(l1)

            l1 = ['']*len(excel_data3[1])
            l1[3] = ('Sem PO = ')
            l1[4] = (sem_PO)
            excel_data3.append(l1)

            Main_excel_sheet['Sem'] = [sem_PO]


            #Sheet4

            for row in worksheet4.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data4.append(row_data)
            
            list04 = ['Assignment','Quiz','Descriptive','CO wise attainment','SEE','Direct CO attainment','Overall CO attainment']
            excel_data4[0].extend(list04)


            assign = Main_excel_sheet['Assignment']
            quiz = Main_excel_sheet['Quiz']
            mid = Main_excel_sheet['Mid']
            sem = Main_excel_sheet['Sem']

            co_wise_attainment = []
            direct_co_attainment = []
            overall_co_attainment = []


            for i in range(1,7):
                temp1 = assign[i-1]
                excel_data4[i].append(assign[i-1])
                if i<4: 
                    temp2 = quiz[0]
                    excel_data4[i].append(quiz[0])
                else: 
                    temp2 = quiz[1]
                    excel_data4[i].append(quiz[1])
                temp3 = mid[i-1]
                excel_data4[i].append(mid[i-1])
                co_wise = round((temp1+temp2+temp3)/3)
                co_wise_attainment.append(co_wise)
                excel_data4[i].append(co_wise)
                excel_data4[i].append(sem[0])
                direct_co = 0.5*co_wise+0.5*sem[0]
                excel_data4[i].append(direct_co)
                direct_co_attainment.append(direct_co)
                overall = round(0.80*direct_co + 0.20*float(excel_data4[i][1]),2)
                excel_data4[i].append(overall)
                overall_co_attainment.append(overall)
            average_co = round(sum(overall_co_attainment)/len(overall_co_attainment),3)
            excel_data4.append(['','','','','','','','Average CO Attainment',average_co])


            #Sheet 5

            for row in worksheet5.iter_rows():
                row_data = list()
                for cell in row:
                    row_data.append(str(cell.value))
                excel_data5.append(row_data) 

            excel_data6 += excel_data5          

            PO1,PO2,PO3,PO4,PO5,PO6,PO7,PO8,PO9,PO10,PO11,PO12,PSO1,PSO2,PSO3 = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
            PO1_II,PO2_II,PO3_II,PO4_II,PO5_II,PO6_II,PO7_II,PO8_II,PO9_II,PO10_II,PO11_II,PO12_II,PSO1_II,PSO2_II,PSO3_II = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
            PO_Average,PSO_Average = [],[]
            POII_Average,PSOII_Average = [],[]
            for i in range(2,len(excel_data5)):
                if none_or_not(excel_data5[i][2]):
                    PO1.append(int(excel_data5[i][2]))
                    PO1_II.append(round((int(excel_data5[i][2])*average_co)/3,2))
                if none_or_not(excel_data5[i][3]):
                    PO2.append(int(excel_data5[i][3]))
                    PO2_II.append(round((int(excel_data5[i][3])*average_co)/3,2))
                if none_or_not(excel_data5[i][4]):
                    PO3.append(int(excel_data5[i][4]))
                    PO3_II.append(round((int(excel_data5[i][4])*average_co)/3,2))
                if none_or_not(excel_data5[i][5]):
                    PO4.append(int(excel_data5[i][5]))
                    PO4_II.append(round((int(excel_data5[i][5])*average_co)/3,2))
                if none_or_not(excel_data5[i][6]):
                    PO5.append(int(excel_data5[i][6]))
                    PO5_II.append(round((int(excel_data5[i][6])*average_co)/3,2))
                if none_or_not(excel_data5[i][7]):
                    PO6.append(int(excel_data5[i][7]))
                    PO6_II.append(round((int(excel_data5[i][7])*average_co)/3,2))
                if none_or_not(excel_data5[i][8]):
                    PO7.append(int(excel_data5[i][8]))
                    PO7_II.append(round((int(excel_data5[i][8])*average_co)/3,2))
                if none_or_not(excel_data5[i][9]):
                    PO8.append(int(excel_data5[i][9]))
                    PO8_II.append(round((int(excel_data5[i][9])*average_co)/3,2))
                if none_or_not(excel_data5[i][10]):
                    PO9.append(int(excel_data5[i][10]))
                    PO9_II.append(round((int(excel_data5[i][10])*average_co)/3,2))
                if none_or_not(excel_data5[i][11]):
                    PO10.append(int(excel_data5[i][11]))
                    PO10_II.append(round((int(excel_data5[i][11])*average_co)/3,2))
                if none_or_not(excel_data5[i][12]):
                    PO11.append(int(excel_data5[i][12]))
                    PO11_II.append(round((int(excel_data5[i][12])*average_co)/3,2))
                if none_or_not(excel_data5[i][13]):
                    PO12.append(int(excel_data5[i][13]))
                    PO12_II.append(round((int(excel_data5[i][13])*average_co)/3,2))
                if none_or_not(excel_data5[i][14]):
                    PSO1.append(int(excel_data5[i][14]))
                    PSO1_II.append(round((int(excel_data5[i][14])*average_co)/3,2))
                if none_or_not(excel_data5[i][15]):
                    PSO2.append(int(excel_data5[i][15]))
                    PSO2_II.append(round((int(excel_data5[i][15])*average_co)/3,2))
                if none_or_not(excel_data5[i][16]):
                    PSO3.append(int(excel_data5[i][16]))
                    PSO3_II.append(round((int(excel_data5[i][16])*average_co)/3,2))
            

            if len(PO1): 
                PO_Average.append(round(sum(PO1)/len(PO1),2))
                POII_Average.append(round(sum(PO1_II)/len(PO1_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)
            
            if len(PO2): 
                PO_Average.append(round(sum(PO2)/len(PO2),2)) 
                POII_Average.append(round(sum(PO2_II)/len(PO2_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO3): 
                PO_Average.append(round(sum(PO3)/len(PO3),2)) 
                POII_Average.append(round(sum(PO3_II)/len(PO3_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO4): 
                PO_Average.append(round(sum(PO4)/len(PO4),2))
                POII_Average.append(round(sum(PO4_II)/len(PO4_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO5): 
                PO_Average.append(round(sum(PO5)/len(PO5),2))
                POII_Average.append(round(sum(PO5_II)/len(PO5_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO6): 
                PO_Average.append(round(sum(PO6)/len(PO6),2)) 
                POII_Average.append(round(sum(PO6_II)/len(PO6_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO7): 
                PO_Average.append(round(sum(PO7)/len(PO7),2)) 
                POII_Average.append(round(sum(PO7_II)/len(PO7_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO8): 
                PO_Average.append(round(sum(PO8)/len(PO8),2)) 
                POII_Average.append(round(sum(PO8_II)/len(PO8_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO9): 
                PO_Average.append(round(sum(PO9)/len(PO9),2))
                POII_Average.append(round(sum(PO9_II)/len(PO9_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO10): 
                PO_Average.append(round(sum(PO10)/len(PO10),2))
                POII_Average.append(round(sum(PO10_II)/len(PO10_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO11): 
                PO_Average.append(round(sum(PO11)/len(PO11),2))
                POII_Average.append(round(sum(PO11_II)/len(PO11_II),2))
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PO12): 
                PO_Average.append(round(sum(PO12)/len(PO12),2))
                POII_Average.append(round(sum(PO12_II)/len(PO12_II),2)) 
            else:
                PO_Average.append(0)
                POII_Average.append(0)

            if len(PSO1): 
                PSO_Average.append(round(sum(PSO1)/len(PSO1),2)) 
                PSOII_Average.append(round(sum(PSO1_II)/len(PSO1_II),2))
            else:
                PSO_Average.append(0)
                PSOII_Average.append(0)

            if len(PSO2): 
                PSO_Average.append(round(sum(PSO2)/len(PSO2),2)) 
                PSOII_Average.append(round(sum(PSO2_II)/len(PSO2_II),2))
            else:
                PSO_Average.append(0)
                PSOII_Average.append(0)

            if len(PSO3): 
                PSO_Average.append(round(sum(PSO3)/len(PSO3),2))
                PSOII_Average.append(round(sum(PSO3_II)/len(PSO3_II),2)) 
            else:
                PSO_Average.append(0)
                PSOII_Average.append(0)

            # if len(PO1): PSO_Average.append(round(sum(PO1)/len(PO1),2)) 
            # else:PSO_Average.append(0)

            # if len(PO1): PSO_Average.append(round(sum(PO1)/len(PO1),2)) 
            # else:PSO_Average.append(0)


            
            # for i in range(2,len())
            excel_data5.append(['','']+PO_Average+PSO_Average)
            excel_data5.append(['','']+POII_Average+PSOII_Average)
            denominator_PO,denominator_PSO = 0,0
            for i in POII_Average:
                if i!=0: denominator_PO+=1
            for i in PSOII_Average:
                if i!=0: denominator_PSO+=1
            Overall_PO_Attainment = round(sum(POII_Average)/denominator_PO,2)
            Overall_PSO_Attainment = round(sum(PSOII_Average)/denominator_PSO,2)
            excel_data5.append([Overall_PO_Attainment,Overall_PSO_Attainment])

            for i in range(len(excel_data5)):
                for j in range(len(excel_data5[i])):
                    if excel_data5[i][j]=='None' or excel_data5[i][j]==0:
                        excel_data5[i][j] = ''


            #Sheet6
            # for row in worksheet6.iter_rows():
            #     row_data = list()
            #     for cell in row:
            #         row_data.append(str(cell.value))
            #     excel_data6.append(row_data)     
            

            S_PO1,S_PO2,S_PO3,S_PO4,S_PO5,S_PO6,S_PO7,S_PO8,S_PO9,S_PO10,S_PO11,S_PO12,S_PSO1,S_PSO2,S_PSO3 = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
            # S_PO1_II,S_PO2_II,S_PO3_II,S_PO4_II,S_PO5_II,S_PO6_II,S_PO7_II,S_PO8_II,S_PO9_II,S_PO10_II,S_PO11_II,S_PO12_II,S_PSO1_II,S_PSO2_II,S_PSO3_II = [],[],[],[],[],[],[],[],[],[],[],[],[],[],[]
            S_PO_Average,S_PSO_Average = [],[]
            # S_POII_Average,S_PSOII_Average = [],[]
            for i in range(2,len(excel_data6)):
                if none_or_not(excel_data6[i][2]):
                    S_PO1.append(int(excel_data6[i][2]))
                if none_or_not(excel_data6[i][3]):
                    S_PO2.append(int(excel_data6[i][3]))
                if none_or_not(excel_data6[i][4]):
                    S_PO3.append(int(excel_data6[i][4]))
                if none_or_not(excel_data6[i][5]):
                    S_PO4.append(int(excel_data6[i][5]))
                if none_or_not(excel_data6[i][6]):
                    S_PO5.append(int(excel_data6[i][6]))
                if none_or_not(excel_data6[i][7]):
                    S_PO6.append(int(excel_data6[i][7]))
                if none_or_not(excel_data6[i][8]):
                    S_PO7.append(int(excel_data6[i][8]))
                if none_or_not(excel_data6[i][9]):
                    S_PO8.append(int(excel_data6[i][9]))
                if none_or_not(excel_data6[i][10]):
                    S_PO9.append(int(excel_data6[i][10]))
                if none_or_not(excel_data6[i][11]):
                    S_PO10.append(int(excel_data6[i][11]))
                if none_or_not(excel_data6[i][12]):
                    S_PO11.append(int(excel_data6[i][12]))
                if none_or_not(excel_data6[i][13]):
                    S_PO12.append(int(excel_data6[i][13]))
                if none_or_not(excel_data6[i][14]):
                    S_PSO1.append(int(excel_data6[i][14]))
                if none_or_not(excel_data6[i][15]):
                    S_PSO2.append(int(excel_data6[i][15]))
                if none_or_not(excel_data6[i][16]):
                    S_PSO3.append(int(excel_data6[i][16]))
                
            

            if len(S_PO1): 
                S_PO_Average.append(round(sum(S_PO1)/len(S_PO1),2))
            else:
                S_PO_Average.append(0)
            
            if len(S_PO2): 
                S_PO_Average.append(round(sum(S_PO2)/len(S_PO2),2)) 
            else:
                S_PO_Average.append(0)

            if len(S_PO3): 
                S_PO_Average.append(round(sum(S_PO3)/len(S_PO3),2)) 
            else:
                S_PO_Average.append(0)

            if len(S_PO4): 
                S_PO_Average.append(round(sum(S_PO4)/len(S_PO4),2))
            else:
                S_PO_Average.append(0)

            if len(S_PO5): 
                S_PO_Average.append(round(sum(S_PO5)/len(S_PO5),2))
            else:
                S_PO_Average.append(0)

            if len(S_PO6): 
                S_PO_Average.append(round(sum(S_PO6)/len(S_PO6),2)) 
            else:
                S_PO_Average.append(0)

            if len(S_PO7): 
                S_PO_Average.append(round(sum(S_PO7)/len(S_PO7),2)) 
            else:
                S_PO_Average.append(0)

            if len(S_PO8): 
                S_PO_Average.append(round(sum(S_PO8)/len(S_PO8),2)) 
            else:
                S_PO_Average.append(0)

            if len(S_PO9): 
                S_PO_Average.append(round(sum(S_PO9)/len(S_PO9),2))
            else:
                S_PO_Average.append(0)

            if len(S_PO10): 
                S_PO_Average.append(round(sum(S_PO10)/len(S_PO10),2))
            else:
                S_PO_Average.append(0)

            if len(S_PO11): 
                S_PO_Average.append(round(sum(S_PO11)/len(S_PO11),2))
            else:
                S_PO_Average.append(0)

            if len(S_PO12): 
                S_PO_Average.append(round(sum(S_PO12)/len(S_PO12),2))
            else:
                S_PO_Average.append(0)

            if len(S_PSO1): 
                S_PSO_Average.append(round(sum(S_PSO1)/len(S_PSO1),2)) 
            else:
                S_PSO_Average.append(0)

            if len(S_PSO2): 
                S_PSO_Average.append(round(sum(S_PSO2)/len(S_PSO2),2)) 
            else:
                S_PSO_Average.append(0)

            if len(S_PSO3): 
                S_PSO_Average.append(round(sum(S_PSO3)/len(S_PSO3),2))
            else:
                S_PSO_Average.append(0)

        


            
            # for i in range(2,len())
            excel_data6.append(['']+S_PO_Average+S_PSO_Average)

            for i in range(len(excel_data6)):
                for j in range(len(excel_data6[i])):
                    if excel_data6[i][j]=='None' or excel_data6[i][j]==0:
                        excel_data6[i][j] = ''
            


            return render(request,'po_output.html', {"depart":depart[Branch],"assign2":assign2,"final":average_co,"over":overall_co_attainment,"direct":direct_co_attainment,"internal":co_wise_attainment,"main":Main_excel_sheet,"excel_data1":excel_data1,"excel_data2":excel_data2,"excel_data3":excel_data3,"excel_data4":excel_data4,"excel_data5":excel_data5,"Details":details,"excel_data6":excel_data6})
            #return render(request,'po_output.html',{"excel_data1":excel_data1,"excel_data2":excel_data2,"excel_data3":excel_data3})